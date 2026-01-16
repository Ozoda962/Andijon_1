# # app/admin.py
# import json
# import datetime
# import os

# from django import forms
# from django.contrib import admin
# from django.urls import path, reverse
# from django.shortcuts import get_object_or_404
# from django.template.response import TemplateResponse
# from django.utils import timezone
# from django.utils.safestring import mark_safe
# from django.utils.translation import gettext_lazy as _
# from django.http import HttpResponse
# from django.conf import settings

# from django.db.models import Avg, Sum
# from django.db.models.functions import TruncDay, TruncMonth

# from openpyxl import load_workbook
# from unfold.admin import ModelAdmin

# from .models import (
#     Direction, Location, Section,
#     DatchikType, Datchik,
#     DatchikFormula, DatchikLog,
#     DataloggerChannel, RawReading,
# )


# # ===================== HELPERS =====================

# def detect_kind_from_title(title: str) -> str:
#     t = (title or "").strip().lower()
#     if t.startswith("sh.d-"):
#         return "shelemer"
#     if t.startswith("v/s"):
#         return "vodosliv"
#     if t.startswith("o.d-"):
#         return "niveller"
#     if "otves" in t:
#         return "otves"
#     parts = t.split("-")
#     if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
#         return "atves"
#     if "byef" in t or "byev" in t:
#         return "byef"
#     return "piezometr"


# def detect_kind_from_type(obj: Datchik) -> str:
#     if getattr(obj, "datchik_type", None) and getattr(obj.datchik_type, "title", None):
#         t = (obj.datchik_type.title or "").strip().lower()
#         if "shelemer" in t or "sh" in t:
#             return "shelemer"
#         if "vodosliv" in t or "v/s" in t:
#             return "vodosliv"
#         if "niv" in t or "tilt" in t or "ogish" in t:
#             return "niveller"
#         if "otves" in t:
#             return "otves"
#         if "atves" in t:
#             return "atves"
#         if "byef" in t or "byev" in t or "yuqori" in t:
#             return "byef"
#         return "piezometr"
#     return detect_kind_from_title(obj.title)


# def iso_date_default(value: str, fallback: datetime.date) -> str:
#     try:
#         if value:
#             datetime.date.fromisoformat(value)
#             return value
#     except Exception:
#         pass
#     return str(fallback)


# def export_logs_to_excel_template(request, queryset, template_path: str, filename: str):
#     if not os.path.exists(template_path):
#         return HttpResponse(f"Template topilmadi: {template_path}", status=404)

#     wb = load_workbook(template_path)
#     ws = wb.active

#     META = {
#         "datchik": "B2",
#         "direction": "B3",
#         "location": "B4",
#         "section": "B5",
#         "date_from": "E3",
#         "date_to": "E4",
#     }

#     DATA_START_ROW = 8

#     COLUMNS = [
#         ("sana", "A"),
#         ("bosim_MPa", "B"),
#         ("bosim_m", "C"),
#         ("temperatura", "D"),
#         ("suv_sathi", "E"),
#         ("suv_sarfi", "F"),
#         ("loyqa", "G"),
#         ("deformatsiya_x", "H"),
#         ("deformatsiya_y", "I"),
#         ("deformatsiya_z", "J"),
#         ("temperatura_x", "K"),
#         ("temperatura_y", "L"),
#         ("temperatura_z", "M"),
#         ("sina", "N"),
#         ("sinb", "O"),
#         ("vektor_ogish_korsatgichi", "P"),
#     ]

#     first = queryset.select_related(
#         "formula", "formula__datchik",
#         "formula__datchik__direction",
#         "formula__datchik__location",
#         "formula__datchik__section",
#     ).first()

#     if first and first.formula and first.formula.datchik:
#         d = first.formula.datchik
#         ws[META["datchik"]] = d.title
#         ws[META["direction"]] = getattr(d.direction, "title", "") if d.direction else ""
#         ws[META["location"]] = getattr(d.location, "title", "") if d.location else ""
#         ws[META["section"]] = getattr(d.section, "title", "") if d.section else ""
#     else:
#         ws[META["datchik"]] = "—"

#     date_from = request.GET.get("sana__date__gte") or request.GET.get("sana__gte") or request.GET.get("from") or ""
#     date_to = request.GET.get("sana__date__lte") or request.GET.get("sana__lte") or request.GET.get("to") or ""
#     if date_from:
#         ws[META["date_from"]] = date_from
#     if date_to:
#         ws[META["date_to"]] = date_to

#     row = DATA_START_ROW
#     for log in queryset.iterator(chunk_size=5000):
#         for field, col in COLUMNS:
#             val = getattr(log, field, None)
#             if field == "sana" and val:
#                 val = timezone.localtime(val).replace(tzinfo=None)
#             ws[f"{col}{row}"] = val
#         row += 1

#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = f'attachment; filename="{filename}"'
#     wb.save(response)
#     return response


# # ===================== FILTERS =====================

# class DatchikKindFilter(admin.SimpleListFilter):
#     title = _("Datchik turi")
#     parameter_name = "kind"

#     def lookups(self, request, model_admin):
#         return (
#             ("piezometr", _("Piezometr")),
#             ("byef", _("Yuqori byef suv sathi")),
#             ("shelemer", _("Shelemer (SH)")),
#             ("vodosliv", _("Vodosliv (V/S)")),
#             ("atves", _("Atves (845-2)")),
#             ("otves", _("Otves")),
#             ("niveller", _("Niveller (O.D-)")),
#         )

#     def queryset(self, request, queryset):
#         val = self.value()
#         if not val:
#             return queryset
#         if val == "shelemer":
#             return queryset.filter(title__istartswith="SH.D-")
#         if val == "vodosliv":
#             return queryset.filter(title__istartswith="V/S")
#         if val == "niveller":
#             return queryset.filter(title__istartswith="O.D-")
#         if val == "otves":
#             return queryset.filter(title__icontains="otves")
#         if val == "atves":
#             return queryset.filter(title__regex=r"^\d+\-\d+$")
#         if val == "byef":
#             return queryset.filter(title__icontains="byef")
#         return queryset


# class DatchikLogKindFilter(admin.SimpleListFilter):
#     title = _("Log turi")
#     parameter_name = "kind"

#     def lookups(self, request, model_admin):
#         return DatchikKindFilter().lookups(request, model_admin)

#     def queryset(self, request, queryset):
#         val = self.value()
#         if not val:
#             return queryset
#         if val == "shelemer":
#             return queryset.filter(formula__datchik__title__istartswith="SH.D-")
#         if val == "vodosliv":
#             return queryset.filter(formula__datchik__title__istartswith="V/S")
#         if val == "niveller":
#             return queryset.filter(formula__datchik__title__istartswith="O.D-")
#         if val == "otves":
#             return queryset.filter(formula__datchik__title__icontains="otves")
#         if val == "atves":
#             return queryset.filter(formula__datchik__title__regex=r"^\d+\-\d+$")
#         if val == "byef":
#             return queryset.filter(formula__datchik__title__icontains="byef")
#         return queryset


# # ===================== INLINES =====================

# class LocationInline(admin.TabularInline):
#     model = Location
#     extra = 0


# class SectionInline(admin.TabularInline):
#     model = Section
#     extra = 0


# class DatchikInline(admin.TabularInline):
#     model = Datchik
#     extra = 0


# class DatchikLogInline(admin.TabularInline):
#     model = DatchikLog
#     extra = 0
#     fk_name = "formula"
#     can_delete = False
#     show_change_link = True
#     readonly_fields = (
#         "sana",
#         "bosim_MPa", "bosim_m", "bosim_sm", "bosim_mm",
#         "suv_sathi", "temperatura", "suv_sarfi", "loyqa",
#         "deformatsiya_x", "deformatsiya_y", "deformatsiya_z",
#         "temperatura_x", "temperatura_y", "temperatura_z",
#         "sina", "sinb", "vektor_ogish_korsatgichi",
#     )


# class DatchikFormulaInline(admin.StackedInline):
#     model = DatchikFormula
#     extra = 0
#     show_change_link = True

#     fields = (
#         "criterion_1", "criterion_2",

#         "bosim_MPa", "bosim_m",
#         "suv_sathi", "temperatura", "suv_sarfi", "loyqa",

#         "deformatsiya_x", "deformatsiya_y", "deformatsiya_z",
#         "temperatura_x", "temperatura_y", "temperatura_z",

#         "sina", "sinb",
#         "vektor_ogish_korsatgichi",
#     )


# # ===================== ADMINS =====================

# @admin.register(Direction)
# class DirectionAdmin(ModelAdmin):
#     list_display = ("id", "title")
#     search_fields = ("title",)
#     inlines = [LocationInline]


# @admin.register(Location)
# class LocationAdmin(ModelAdmin):
#     list_display = ("id", "code", "title", "direction")
#     list_filter = ("direction",)
#     search_fields = ("title", "code")
#     inlines = [SectionInline, DatchikInline]


# @admin.register(Section)
# class SectionAdmin(ModelAdmin):
#     list_display = ("id", "title", "location", "direction")
#     list_filter = ("direction", "location")
#     search_fields = ("title",)
#     inlines = [DatchikInline]


# @admin.register(DatchikType)
# class DatchikTypeAdmin(ModelAdmin):
#     list_display = ("id", "title", "interval_minutes", "per_day")
#     search_fields = ("title",)


# @admin.register(Datchik)
# class DatchikAdmin(ModelAdmin):
#     list_display = (
#         "id", "title", "kind",
#         "direction", "location", "section", "datchik_type",
#         "chart_link",
#     )
#     list_filter = (DatchikKindFilter, "direction", "location", "section", "datchik_type")
#     search_fields = ("title",)
#     autocomplete_fields = ("direction", "location", "section", "datchik_type")
#     list_select_related = ("direction", "location", "section", "datchik_type")

#     inlines = [DatchikFormulaInline]

#     @admin.display(description=_("Turi"))
#     def kind(self, obj):
#         return detect_kind_from_type(obj)

#     @admin.display(description=_("Grafik"))
#     def chart_link(self, obj):
#         url = reverse("admin:app_datchik_chart", args=[obj.id])
#         return mark_safe(f'<a class="button" href="{url}">📈 Grafik</a>')

#     def get_urls(self):
#         urls = super().get_urls()
#         custom = [
#             path(
#                 "<int:datchik_id>/chart/",
#                 self.admin_site.admin_view(self.chart_view),
#                 name="app_datchik_chart",
#             ),
#         ]
#         return custom + urls

#     def chart_view(self, request, datchik_id: int):
#         datchik = get_object_or_404(Datchik, pk=datchik_id)
#         today = timezone.localdate()

#         date_from = iso_date_default(request.GET.get("from"), today - datetime.timedelta(days=7))
#         date_to = iso_date_default(request.GET.get("to"), today)

#         chart_type = (request.GET.get("chart") or "line").strip().lower()
#         period = (request.GET.get("period") or "raw").strip().lower()
#         agg = (request.GET.get("agg") or "avg").strip().lower()

#         if chart_type not in ("line", "bar"):
#             chart_type = "line"
#         if period not in ("raw", "day", "month"):
#             period = "raw"
#         if agg not in ("avg", "sum"):
#             agg = "avg"

#         tz = timezone.get_current_timezone()
#         dt_from = timezone.make_aware(datetime.datetime.fromisoformat(date_from + " 00:00:00"), tz)
#         dt_to = timezone.make_aware(datetime.datetime.fromisoformat(date_to + " 23:59:59"), tz)

#         qs = (
#             DatchikLog.objects
#             .filter(formula__datchik=datchik, sana__range=(dt_from, dt_to))
#             .order_by("sana")
#         )

#         kind = detect_kind_from_type(datchik)

#         if kind == "shelemer":
#             fields = [
#                 ("deformatsiya_x", "Def X"),
#                 ("temperatura_x", "Temp X"),
#                 ("deformatsiya_y", "Def Y"),
#                 ("temperatura_y", "Temp Y"),
#                 ("deformatsiya_z", "Def Z"),
#                 ("temperatura_z", "Temp Z"),
#             ]
#         elif kind == "vodosliv":
#             fields = [("suv_sarfi", "Suv sarfi"), ("loyqa", "Loyqa")]
#         elif kind in ("otves", "atves"):
#             fields = [("deformatsiya_x", "Def X"), ("deformatsiya_y", "Def Y"), ("temperatura", "Temp")]
#         elif kind == "niveller":
#             fields = [("sina", "Sin(a)"), ("sinb", "Sin(b)"), ("temperatura", "Temp")]
#         elif kind == "byef":
#             # BYEF: vektor og'ish X/Y dan (logda saqlangan bo'lsa ham chizamiz)
#             fields = [("deformatsiya_x", "Def X"), ("deformatsiya_y", "Def Y"), ("vektor_ogish_korsatgichi", "Vektor og'ish")]
#         else:
#             fields = [("bosim_MPa", "Bosim (MPa)"), ("bosim_m", "Bosim (m)"), ("suv_sathi", "Suv sathi"), ("temperatura", "Temp")]

#         labels = []
#         datasets = []

#         if period == "raw":
#             series = {f[0]: [] for f in fields}
#             for log in qs:
#                 labels.append(timezone.localtime(log.sana).strftime("%Y-%m-%d %H:%M"))
#                 for field, _label in fields:
#                     series[field].append(getattr(log, field, None))
#             datasets = [{"label": label, "data": series[field]} for field, label in fields]
#         else:
#             if period == "day":
#                 trunc = TruncDay("sana")
#                 label_fmt = "%Y-%m-%d"
#             else:
#                 trunc = TruncMonth("sana")
#                 label_fmt = "%Y-%m"

#             base = qs.annotate(t=trunc).values("t").distinct().order_by("t")
#             label_map = [row["t"] for row in base]
#             labels = [timezone.localtime(t).strftime(label_fmt) for t in label_map]

#             for field, human_label in fields:
#                 agg_expr = Sum(field) if agg == "sum" else Avg(field)
#                 rows = (
#                     qs.annotate(t=trunc)
#                       .values("t")
#                       .annotate(v=agg_expr)
#                       .order_by("t")
#                 )
#                 vmap = {r["t"]: (float(r["v"]) if r["v"] is not None else None) for r in rows}
#                 data = [vmap.get(t, None) for t in label_map]
#                 datasets.append({"label": human_label, "data": data})

#         ctx = dict(
#             self.admin_site.each_context(request),
#             title=f"Grafik: {datchik.title}",
#             datchik=datchik,
#             date_from=date_from,
#             date_to=date_to,
#             chart_type=chart_type,
#             period=period,
#             agg=agg,
#             labels_json=json.dumps(labels, ensure_ascii=False),
#             datasets_json=json.dumps(datasets, ensure_ascii=False),
#         )
#         return TemplateResponse(request, "admin/app/datchik_chart.html", ctx)


# @admin.register(DatchikFormula)
# class DatchikFormulaAdmin(ModelAdmin):
#     list_display = ("id", "datchik", "direction", "location", "section")
#     list_filter = ("datchik__direction", "datchik__location", "datchik__section")
#     search_fields = ("datchik__title",)
#     autocomplete_fields = ("datchik",)
#     list_select_related = ("datchik", "datchik__direction", "datchik__location", "datchik__section")
#     inlines = [DatchikLogInline]

#     # hammasi ochiq bo‘lsin + criterionlar ham doim
#     fields = (
#         "datchik",
#         "criterion_1", "criterion_2",

#         "bosim_MPa", "bosim_m",
#         "suv_sathi", "temperatura", "suv_sarfi", "loyqa",

#         "deformatsiya_x", "deformatsiya_y", "deformatsiya_z",
#         "temperatura_x", "temperatura_y", "temperatura_z",

#         "sina", "sinb",
#         "vektor_ogish_korsatgichi",
#     )

#     @admin.display(description=_("Direction"))
#     def direction(self, obj):
#         return obj.datchik.direction if obj.datchik else "-"

#     @admin.display(description=_("Location"))
#     def location(self, obj):
#         return obj.datchik.location if obj.datchik else "-"

#     @admin.display(description=_("Section"))
#     def section(self, obj):
#         return obj.datchik.section if obj.datchik else "-"


# @admin.register(DatchikLog)
# class DatchikLogAdmin(ModelAdmin):
#     change_list_template = "admin/app/datchiklog/change_list_with_export.html"

#     list_display = (
#         "id", "datchik_title", "datchik_kind", "sana",
#         "bosim_MPa", "bosim_m", "temperatura", "suv_sathi", "suv_sarfi", "loyqa",
#         "deformatsiya_x", "deformatsiya_y", "deformatsiya_z",
#         "temperatura_x", "temperatura_y", "temperatura_z",
#         "sina", "sinb", "vektor_ogish_korsatgichi",
#     )
#     list_filter = (
#         DatchikLogKindFilter,
#         "sana",
#         "formula__datchik__direction",
#         "formula__datchik__location",
#         "formula__datchik__section",
#     )
#     search_fields = ("formula__datchik__title",)
#     date_hierarchy = "sana"
#     readonly_fields = ("sana",)
#     list_select_related = (
#         "formula", "formula__datchik",
#         "formula__datchik__direction",
#         "formula__datchik__location",
#         "formula__datchik__section",
#     )
#     ordering = ("-sana",)
#     actions = []

#     def get_urls(self):
#         urls = super().get_urls()
#         custom = [
#             path(
#                 "export-excel/",
#                 self.admin_site.admin_view(self.export_excel_view),
#                 name="app_datchiklog_export_excel",
#             ),
#         ]
#         return custom + urls

#     def export_excel_view(self, request):
#         cl = self.get_changelist_instance(request)
#         qs = cl.get_queryset(request)

#         template_path = os.path.join(
#             settings.BASE_DIR,
#             "app",
#             "resources",
#             "excel_templates",
#             "export_template.xlsx",   # o‘zingizdagi template nomi
#         )

#         return export_logs_to_excel_template(
#             request=request,
#             queryset=qs,
#             template_path=template_path,
#             filename="export_logs.xlsx",
#         )

#     @admin.display(description=_("Datchik"))
#     def datchik_title(self, obj):
#         return obj.formula.datchik.title if obj.formula and obj.formula.datchik else "-"

#     @admin.display(description=_("Turi"))
#     def datchik_kind(self, obj):
#         if obj.formula and obj.formula.datchik:
#             return detect_kind_from_type(obj.formula.datchik)
#         return "-"


# @admin.register(DataloggerChannel)
# class DataloggerChannelAdmin(ModelAdmin):
#     list_display = ("node_id", "channel", "value_type", "datchik", "direction", "location", "section")
#     list_filter = ("node_id", "channel", "value_type", "datchik__direction", "datchik__location", "datchik__section")
#     search_fields = ("node_id", "channel", "datchik__title")
#     autocomplete_fields = ("datchik",)
#     ordering = ("node_id", "channel")
#     list_per_page = 50
#     list_select_related = ("datchik", "datchik__direction", "datchik__location", "datchik__section")

#     @admin.display(description="Direction")
#     def direction(self, obj):
#         return obj.datchik.direction if obj.datchik else "-"

#     @admin.display(description="Location")
#     def location(self, obj):
#         return obj.datchik.location if obj.datchik else "-"

#     @admin.display(description="Section")
#     def section(self, obj):
#         return obj.datchik.section if obj.datchik else "-"


# @admin.register(RawReading)
# class RawReadingAdmin(ModelAdmin):
#     list_display = ("id", "datchik", "ts", "value_type", "raw_value", "source_file", "created_at")
#     list_filter = ("value_type", "ts", "datchik__direction", "datchik__location", "datchik__section")
#     search_fields = ("datchik__title", "source_file")
#     autocomplete_fields = ("datchik",)
#     date_hierarchy = "ts"
#     ordering = ("-ts",)
#     list_select_related = ("datchik", "datchik__direction", "datchik__location", "datchik__section")



# app/admin.py
import json
import datetime
import os

from django.contrib import admin
from django.urls import path, reverse
from django.shortcuts import get_object_or_404
from django.template.response import TemplateResponse
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse
from django.conf import settings

from django.db.models import Avg, Sum
from django.db.models.functions import TruncDay, TruncMonth

from openpyxl import load_workbook
from unfold.admin import ModelAdmin

from .models import (
    Direction, Location, Section,
    DatchikType, Datchik,
    DatchikFormula, DatchikLog,
    DataloggerChannel, RawReading,
)

# ===================== HELPERS =====================

def detect_kind_from_title(title: str) -> str:
    t = (title or "").strip().lower()
    if t.startswith("sh.d-"):
        return "shelemer"
    if t.startswith("v/s"):
        return "vodosliv"
    if t.startswith("o.d-"):
        return "niveller"
    if "otves" in t:
        return "otves"
    parts = t.split("-")
    if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
        return "atves"
    if "byef" in t or "byev" in t:
        return "byef"
    return "piezometr"


def detect_kind_from_type(obj: Datchik) -> str:
    if getattr(obj, "datchik_type", None) and getattr(obj.datchik_type, "title", None):
        t = (obj.datchik_type.title or "").strip().lower()
        if "shelemer" in t or "sh" in t:
            return "shelemer"
        if "vodosliv" in t or "v/s" in t:
            return "vodosliv"
        if "niv" in t or "tilt" in t or "ogish" in t:
            return "niveller"
        if "otves" in t:
            return "otves"
        if "atves" in t:
            return "atves"
        if "byef" in t or "byev" in t or "yuqori" in t:
            return "byef"
        return "piezometr"
    return detect_kind_from_title(obj.title)


def iso_date_default(value: str, fallback: datetime.date) -> str:
    try:
        if value:
            datetime.date.fromisoformat(value)
            return value
    except Exception:
        pass
    return str(fallback)


def export_logs_to_excel_template(request, queryset, template_path: str, filename: str):
    if not os.path.exists(template_path):
        return HttpResponse(f"Template topilmadi: {template_path}", status=404)

    wb = load_workbook(template_path)
    ws = wb.active

    META = {
        "datchik": "B2",
        "direction": "B3",
        "location": "B4",
        "section": "B5",
        "date_from": "E3",
        "date_to": "E4",
    }

    DATA_START_ROW = 8

    COLUMNS = [
        ("sana", "A"),
        ("bosim_MPa", "B"),
        ("bosim_m", "C"),
        ("temperatura", "D"),
        ("suv_sathi", "E"),
        ("suv_sarfi", "F"),
        ("loyqa", "G"),
        ("deformatsiya_x", "H"),
        ("deformatsiya_y", "I"),
        ("deformatsiya_z", "J"),
        ("temperatura_x", "K"),
        ("temperatura_y", "L"),
        ("temperatura_z", "M"),
        ("sina", "N"),
        ("sinb", "O"),
        ("vektor_ogish_korsatgichi", "P"),
    ]

    first = queryset.select_related(
        "formula", "formula__datchik",
        "formula__datchik__direction",
        "formula__datchik__location",
        "formula__datchik__section",
    ).first()

    if first and first.formula and first.formula.datchik:
        d = first.formula.datchik
        ws[META["datchik"]] = d.title
        ws[META["direction"]] = getattr(d.direction, "title", "") if d.direction else ""
        ws[META["location"]] = getattr(d.location, "title", "") if d.location else ""
        ws[META["section"]] = getattr(d.section, "title", "") if d.section else ""
    else:
        ws[META["datchik"]] = "—"

    date_from = request.GET.get("sana__date__gte") or request.GET.get("sana__gte") or request.GET.get("from") or ""
    date_to = request.GET.get("sana__date__lte") or request.GET.get("sana__lte") or request.GET.get("to") or ""
    if date_from:
        ws[META["date_from"]] = date_from
    if date_to:
        ws[META["date_to"]] = date_to

    row = DATA_START_ROW
    for log in queryset.iterator(chunk_size=5000):
        for field, col in COLUMNS:
            val = getattr(log, field, None)
            if field == "sana" and val:
                val = timezone.localtime(val).replace(tzinfo=None)
            ws[f"{col}{row}"] = val
        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ===================== FILTERS =====================

class DatchikKindFilter(admin.SimpleListFilter):
    title = _("Datchik turi")
    parameter_name = "kind"

    def lookups(self, request, model_admin):
        return (
            ("piezometr", _("Piezometr")),
            ("byef", _("Yuqori byef suv sathi")),
            ("shelemer", _("Shelemer (SH)")),
            ("vodosliv", _("Vodosliv (V/S)")),
            ("atves", _("Atves (845-2)")),
            ("otves", _("Otves")),
            ("niveller", _("Niveller (O.D-)")),
        )

    def queryset(self, request, queryset):
        val = self.value()
        if not val:
            return queryset
        if val == "shelemer":
            return queryset.filter(title__istartswith="SH.D-")
        if val == "vodosliv":
            return queryset.filter(title__istartswith="V/S")
        if val == "niveller":
            return queryset.filter(title__istartswith="O.D-")
        if val == "otves":
            return queryset.filter(title__icontains="otves")
        if val == "atves":
            return queryset.filter(title__regex=r"^\d+\-\d+$")
        if val == "byef":
            return queryset.filter(title__icontains="byef")
        return queryset


class DatchikLogKindFilter(admin.SimpleListFilter):
    title = _("Log turi")
    parameter_name = "kind"

    def lookups(self, request, model_admin):
        # SimpleListFilter'ni instantiate qilmaymiz!
        return (
            ("piezometr", _("Piezometr")),
            ("byef", _("Yuqori byef suv sathi")),
            ("shelemer", _("Shelemer (SH)")),
            ("vodosliv", _("Vodosliv (V/S)")),
            ("atves", _("Atves (845-2)")),
            ("otves", _("Otves")),
            ("niveller", _("Niveller (O.D-)")),
        )

    def queryset(self, request, queryset):
        val = self.value()
        if not val:
            return queryset
        if val == "shelemer":
            return queryset.filter(formula__datchik__title__istartswith="SH.D-")
        if val == "vodosliv":
            return queryset.filter(formula__datchik__title__istartswith="V/S")
        if val == "niveller":
            return queryset.filter(formula__datchik__title__istartswith="O.D-")
        if val == "otves":
            return queryset.filter(formula__datchik__title__icontains="otves")
        if val == "atves":
            return queryset.filter(formula__datchik__title__regex=r"^\d+\-\d+$")
        if val == "byef":
            return queryset.filter(formula__datchik__title__icontains="byef")
        return queryset

# ===================== INLINES =====================

class LocationInline(admin.TabularInline):
    model = Location
    extra = 0


class SectionInline(admin.TabularInline):
    model = Section
    extra = 0


class DatchikInline(admin.TabularInline):
    model = Datchik
    extra = 0


class DataloggerChannelInline(admin.TabularInline):
    model = DataloggerChannel
    extra = 0
    show_change_link = True


class DatchikLogInline(admin.TabularInline):
    model = DatchikLog
    extra = 0
    fk_name = "formula"
    can_delete = False
    show_change_link = True
    readonly_fields = (
        "sana",
        "bosim_MPa", "bosim_m", "bosim_sm", "bosim_mm",
        "suv_sathi", "temperatura", "suv_sarfi", "loyqa",
        "deformatsiya_x", "deformatsiya_y", "deformatsiya_z",
        "temperatura_x", "temperatura_y", "temperatura_z",
        "sina", "sinb", "vektor_ogish_korsatgichi",
    )


class DatchikFormulaInline(admin.StackedInline):
    """
    MUHIM: bu inline hech narsa yashirmaydi.
    criterion_1/2 ham, hamma formula maydonlari ham doim ochiq.
    """
    model = DatchikFormula
    extra = 1
    show_change_link = True

    fields = (
        "criterion_1", "criterion_2",

        "bosim_MPa", "bosim_m",
        "suv_sathi", "temperatura", "suv_sarfi", "loyqa",

        "deformatsiya_x", "deformatsiya_y", "deformatsiya_z",
        "temperatura_x", "temperatura_y", "temperatura_z",

        "sina", "sinb",
        "vektor_ogish_korsatgichi",
    )


# ===================== ADMINS =====================

@admin.register(Direction)
class DirectionAdmin(ModelAdmin):
    list_display = ("id", "title")
    search_fields = ("title",)
    inlines = [LocationInline]


@admin.register(Location)
class LocationAdmin(ModelAdmin):
    list_display = ("id", "code", "title", "direction")
    list_filter = ("direction",)
    search_fields = ("title", "code")
    inlines = [SectionInline, DatchikInline]


@admin.register(Section)
class SectionAdmin(ModelAdmin):
    list_display = ("id", "title", "location", "direction")
    list_filter = ("direction", "location")
    search_fields = ("title",)
    inlines = [DatchikInline]


@admin.register(DatchikType)
class DatchikTypeAdmin(ModelAdmin):
    list_display = ("id", "title")
    search_fields = ("title",)


@admin.register(Datchik)
class DatchikAdmin(ModelAdmin):
    list_display = (
        "id", "title", "kind",
        "direction", "location", "section", "datchik_type",
        "chart_link",
    )
    list_filter = (DatchikKindFilter, "direction", "location", "section", "datchik_type")
    search_fields = ("title",)
    autocomplete_fields = ("direction", "location", "section", "datchik_type")
    list_select_related = ("direction", "location", "section", "datchik_type")

    # ✅ DataloggerChannel ham ko'rinadi + Formula hammasi ochiq
    inlines = [DataloggerChannelInline, DatchikFormulaInline]

    @admin.display(description=_("Turi"))
    def kind(self, obj):
        return detect_kind_from_type(obj)

    @admin.display(description=_("Grafik"))
    def chart_link(self, obj):
        url = reverse("admin:app_datchik_chart", args=[obj.id])
        return mark_safe(f'<a class="button" href="{url}">📈 Grafik</a>')

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "<int:datchik_id>/chart/",
                self.admin_site.admin_view(self.chart_view),
                name="app_datchik_chart",
            ),
        ]
        return custom + urls

    def chart_view(self, request, datchik_id: int):
        datchik = get_object_or_404(Datchik, pk=datchik_id)
        today = timezone.localdate()

        date_from = iso_date_default(request.GET.get("from"), today - datetime.timedelta(days=7))
        date_to = iso_date_default(request.GET.get("to"), today)

        chart_type = (request.GET.get("chart") or "line").strip().lower()
        period = (request.GET.get("period") or "raw").strip().lower()
        agg = (request.GET.get("agg") or "avg").strip().lower()

        if chart_type not in ("line", "bar"):
            chart_type = "line"
        if period not in ("raw", "day", "month"):
            period = "raw"
        if agg not in ("avg", "sum"):
            agg = "avg"

        tz = timezone.get_current_timezone()
        dt_from = timezone.make_aware(datetime.datetime.fromisoformat(date_from + " 00:00:00"), tz)
        dt_to = timezone.make_aware(datetime.datetime.fromisoformat(date_to + " 23:59:59"), tz)

        qs = (
            DatchikLog.objects
            .filter(formula__datchik=datchik, sana__range=(dt_from, dt_to))
            .order_by("sana")
        )

        kind = detect_kind_from_type(datchik)

        if kind == "shelemer":
            fields = [
                ("deformatsiya_x", "Def X"),
                ("temperatura_x", "Temp X"),
                ("deformatsiya_y", "Def Y"),
                ("temperatura_y", "Temp Y"),
                ("deformatsiya_z", "Def Z"),
                ("temperatura_z", "Temp Z"),
            ]
        elif kind == "vodosliv":
            fields = [("suv_sarfi", "Suv sarfi"), ("loyqa", "Loyqa")]
        elif kind in ("otves", "atves"):
            fields = [("deformatsiya_x", "Def X"), ("deformatsiya_y", "Def Y"), ("temperatura", "Temp")]
        elif kind == "niveller":
            fields = [("sina", "Sin(a)"), ("sinb", "Sin(b)"), ("temperatura", "Temp")]
        elif kind == "byef":
            fields = [("deformatsiya_x", "Def X"), ("deformatsiya_y", "Def Y"), ("vektor_ogish_korsatgichi", "Vektor og'ish")]
        else:
            fields = [("bosim_MPa", "Bosim (MPa)"), ("bosim_m", "Bosim (m)"), ("suv_sathi", "Suv sathi"), ("temperatura", "Temp")]

        labels = []
        datasets = []

        if period == "raw":
            series = {f[0]: [] for f in fields}
            for log in qs:
                labels.append(timezone.localtime(log.sana).strftime("%Y-%m-%d %H:%M"))
                for field, _label in fields:
                    series[field].append(getattr(log, field, None))
            datasets = [{"label": label, "data": series[field]} for field, label in fields]
        else:
            if period == "day":
                trunc = TruncDay("sana")
                label_fmt = "%Y-%m-%d"
            else:
                trunc = TruncMonth("sana")
                label_fmt = "%Y-%m"

            base = qs.annotate(t=trunc).values("t").distinct().order_by("t")
            label_map = [row["t"] for row in base]
            labels = [timezone.localtime(t).strftime(label_fmt) for t in label_map]

            for field, human_label in fields:
                agg_expr = Sum(field) if agg == "sum" else Avg(field)
                rows = (
                    qs.annotate(t=trunc)
                      .values("t")
                      .annotate(v=agg_expr)
                      .order_by("t")
                )
                vmap = {r["t"]: (float(r["v"]) if r["v"] is not None else None) for r in rows}
                data = [vmap.get(t, None) for t in label_map]
                datasets.append({"label": human_label, "data": data})

        ctx = dict(
            self.admin_site.each_context(request),
            title=f"Grafik: {datchik.title}",
            datchik=datchik,
            date_from=date_from,
            date_to=date_to,
            chart_type=chart_type,
            period=period,
            agg=agg,
            labels_json=json.dumps(labels, ensure_ascii=False),
            datasets_json=json.dumps(datasets, ensure_ascii=False),
        )
        return TemplateResponse(request, "admin/app/datchik_chart.html", ctx)


@admin.register(DatchikFormula)
class DatchikFormulaAdmin(ModelAdmin):
    list_display = ("id", "datchik", "criterion_1", "criterion_2", "direction", "location", "section")
    list_filter = ("datchik__direction", "datchik__location", "datchik__section")
    search_fields = ("datchik__title",)
    autocomplete_fields = ("datchik",)
    list_select_related = ("datchik", "datchik__direction", "datchik__location", "datchik__section")
    inlines = [DatchikLogInline]

    # ✅ hammasi ochiq + criterionlar doim
    fields = (
        "datchik",
        "criterion_1", "criterion_2",

        "bosim_MPa", "bosim_m",
        "suv_sathi", "temperatura", "suv_sarfi", "loyqa",

        "deformatsiya_x", "deformatsiya_y", "deformatsiya_z",
        "temperatura_x", "temperatura_y", "temperatura_z",

        "sina", "sinb",
        "vektor_ogish_korsatgichi",
    )

    @admin.display(description=_("Direction"))
    def direction(self, obj):
        return obj.datchik.direction if obj.datchik else "-"

    @admin.display(description=_("Location"))
    def location(self, obj):
        return obj.datchik.location if obj.datchik else "-"

    @admin.display(description=_("Section"))
    def section(self, obj):
        return obj.datchik.section if obj.datchik else "-"


@admin.register(DatchikLog)
class DatchikLogAdmin(ModelAdmin):
    change_list_template = "admin/app/datchiklog/change_list_with_export.html"

    list_display = (
        "id", "datchik_title", "datchik_kind", "sana",
        "bosim_MPa", "bosim_m", "temperatura", "suv_sathi", "suv_sarfi", "loyqa",
        "deformatsiya_x", "deformatsiya_y", "deformatsiya_z",
        "temperatura_x", "temperatura_y", "temperatura_z",
        "sina", "sinb", "vektor_ogish_korsatgichi",
    )
    list_filter = (
        DatchikLogKindFilter,
        "sana",
        "formula__datchik__direction",
        "formula__datchik__location",
        "formula__datchik__section",
    )
    search_fields = ("formula__datchik__title",)
    date_hierarchy = "sana"
    readonly_fields = ("sana",)
    list_select_related = (
        "formula", "formula__datchik",
        "formula__datchik__direction",
        "formula__datchik__location",
        "formula__datchik__section",
    )
    ordering = ("-sana",)
    actions = []

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "export-excel/",
                self.admin_site.admin_view(self.export_excel_view),
                name="app_datchiklog_export_excel",
            ),
        ]
        return custom + urls

    def export_excel_view(self, request):
        cl = self.get_changelist_instance(request)
        qs = cl.get_queryset(request)

        template_path = os.path.join(
            settings.BASE_DIR,
            "app",
            "resources",
            "excel_templates",
            "export_template.xlsx",
        )

        return export_logs_to_excel_template(
            request=request,
            queryset=qs,
            template_path=template_path,
            filename="export_logs.xlsx",
        )

    @admin.display(description=_("Datchik"))
    def datchik_title(self, obj):
        return obj.formula.datchik.title if obj.formula and obj.formula.datchik else "-"

    @admin.display(description=_("Turi"))
    def datchik_kind(self, obj):
        if obj.formula and obj.formula.datchik:
            return detect_kind_from_type(obj.formula.datchik)
        return "-"


@admin.register(DataloggerChannel)
class DataloggerChannelAdmin(ModelAdmin):
    list_display = ("node_id", "channel", "value_type", "datchik", "direction", "location", "section")
    list_filter = ("node_id", "channel", "value_type", "datchik__direction", "datchik__location", "datchik__section")
    search_fields = ("node_id", "channel", "datchik__title")
    autocomplete_fields = ("datchik",)
    ordering = ("node_id", "channel")
    list_per_page = 50
    list_select_related = ("datchik", "datchik__direction", "datchik__location", "datchik__section")

    @admin.display(description="Direction")
    def direction(self, obj):
        return obj.datchik.direction if obj.datchik else "-"

    @admin.display(description="Location")
    def location(self, obj):
        return obj.datchik.location if obj.datchik else "-"

    @admin.display(description="Section")
    def section(self, obj):
        return obj.datchik.section if obj.datchik else "-"


@admin.register(RawReading)
class RawReadingAdmin(ModelAdmin):
    # ⚠️ Agar RawReading modelda created_at bo'lmasa, list_display dan olib tashlang.
    list_display = ("id", "datchik", "ts", "value_type", "raw_value", "source_file")
    list_filter = ("value_type", "ts", "datchik__direction", "datchik__location", "datchik__section")
    search_fields = ("datchik__title", "source_file")
    autocomplete_fields = ("datchik",)
    date_hierarchy = "ts"
    ordering = ("-ts",)
    list_select_related = ("datchik", "datchik__direction", "datchik__location", "datchik__section")


from django.contrib.auth.models import Group
from django.contrib.auth.admin import GroupAdmin
from django.contrib import admin

admin.site.unregister(Group)

@admin.register(Group)
class CustomGroupAdmin(GroupAdmin):
    def get_model_perms(self, request):
        return super().get_model_perms(request)

from django_celery_beat.models import (
    PeriodicTask,
    CrontabSchedule,
    IntervalSchedule,
    SolarSchedule,
    ClockedSchedule,
)

PeriodicTask._meta.verbose_name = "Davriy vazifa"
PeriodicTask._meta.verbose_name_plural = "Davriy vazifalar"

CrontabSchedule._meta.verbose_name = "Crontab jadvali"
CrontabSchedule._meta.verbose_name_plural = "Crontab jadvallari"

IntervalSchedule._meta.verbose_name = "Interval"
IntervalSchedule._meta.verbose_name_plural = "Intervallar"

SolarSchedule._meta.verbose_name = "Astronomik hodisa"
SolarSchedule._meta.verbose_name_plural = "Astronomik hodisalar"

ClockedSchedule._meta.verbose_name = "Aniq vaqt"
ClockedSchedule._meta.verbose_name_plural = "Aniq vaqtlar"
